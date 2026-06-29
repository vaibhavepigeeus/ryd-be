from datetime import datetime 
import os
from django.conf import settings
from django.core.exceptions import ValidationError
from django.forms import DateField, DateTimeField
from django.utils.timezone import now
import boto3
from .models import FileManagement
import json
from .serializers import FileManagementSerializer
from django.db.models.functions import Cast

from rest_framework import status, permissions
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from .models import FileManagement
# from .utils import handle_file_upload
from django.conf import settings
from django.core.exceptions import ValidationError
from knox.auth import TokenAuthentication
from decouple import config
from users.models import Users
from openpyxl.styles import PatternFill
from io import BytesIO
from openpyxl import load_workbook

# from .helper import upload_file
import requests
import urllib

import csv
import pandas as pd
import logging
logger = logging.getLogger('filemanagement')

AWS_S3_PROCESSED_BUCKET = config('AWS_S3_PROCESSED_BUCKET', default='')

def upload_file_to_s3(file,full_file_path, bucketname, file_name):
    """
    Uploads a file to S3 with dynamic extension support.

    :param file: File object to upload.
    :param bucketname: S3 bucket name where the file should be uploaded.
    :return: S3 file URL of the uploaded file.
    """
    try:
        # s3_client = boto3.client('s3')
        
        s3_client = boto3.client(
            's3',
            aws_access_key_id=config("AWS_ACCESS_KEY"),
            aws_secret_access_key=config("AWS_SECRET_KEY"),
            region_name=config("AWS_S3_REGION_NAME"),
            )
        # s3_client.upload_file(full_file_path, bucketname, file_name)
        s3_client.upload_file(full_file_path, bucketname, file_name)
        # s3_file_url = "https://{}.s3.amazonaws.com/{}".format(bucketname, file.name)
        s3_file_url = "https://{}.s3.amazonaws.com/{}".format(bucketname, file_name)

        return s3_file_url
    
    except Exception as e:
        # Handle exceptions such as errors in file upload, invalid file, etc.
        print(f"Error uploading file to S3: {str(e)}")
        raise e


# Function to count records in a file (CSV or Excel)
def count_records_in_file(file):
    """
    Count the number of records in a file (CSV or Excel).

    :param file: File object to count records from.
    :return: Number of records (rows) in the file.
    """
    file_extension = os.path.splitext(file.name)[1].lower()

    try:
        if file_extension == '.csv':
            # Count records in a CSV file
            file.seek(0)  # Reset file pointer to the beginning
            reader = csv.reader(file)
            record_count = sum(1 for row in reader)  # Count rows in CSV
        elif file_extension in ['.xlsx', '.xls']:
            # Count records in an Excel file using pandas
            file.seek(0)  # Reset file pointer to the beginning
            df = pd.read_excel(file)
            record_count = len(df)  # Number of rows in the DataFrame
        else:
            return 0  # Unknown file type, return 0

        return record_count

    except Exception as e:
        print(f"Error counting records in file: {str(e)}")
        return 0




class FileListAPIView(APIView):
    """
    API View to list all uploaded files.
    """
    # permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        queryset = FileManagement.objects.all()

        # queryset = super().get_queryset()  # Start with the base queryset

        # Get filter parameters from the request
        id = self.request.query_params.get('id', None)
        file_name = self.request.query_params.get('file_name', None)
        module_name = self.request.query_params.get('module_name', None)
        uploaded_by = self.request.query_params.get('uploaded_by', None)
        uploaded_at = self.request.query_params.get('uploaded_at', None)
        from_date = self.request.query_params.get('from_date', None)
        to_date = self.request.query_params.get('to_date', None)

        page_number = int(request.GET.get("skip", 0))
        pageSize = int(request.GET.get("pageSize", 20))

        skip = page_number * pageSize
        
        queryset = queryset.filter(archived=False)
        
        # Apply month, year and id filters if they are provided in the query parameters
        if id:
            queryset = queryset.filter(id=id)

        if file_name:
            queryset = queryset.filter(file_name__icontains=file_name)
        
        if module_name:
            queryset = queryset.filter(module_name__icontains=module_name)

        if uploaded_by:
            queryset = queryset.filter(created_by=uploaded_by)

        if uploaded_at:
            queryset = queryset.filter(created_at=uploaded_at)

        if from_date:
            start_date = datetime.strptime(from_date, "%Y-%m-%d")
            start_datetime = datetime.combine(start_date, datetime.min.time())  # Start of the start date
            queryset = queryset.filter(created_at__gte=start_datetime)            

        if to_date:
            to_date = datetime.strptime(to_date, "%Y-%m-%d")
            to_datetime = datetime.combine(to_date, datetime.max.time())  # Start of the start date
            queryset = queryset.filter(created_at__lte=to_datetime)

        queryset = queryset.order_by('-id')

        serializer = FileManagementSerializer(queryset[skip: skip + pageSize], many=True)
        return Response({'data': serializer.data, 'count': len(queryset)},
                                status=status.HTTP_200_OK)
        
def get_user(request):
    user_id = request.headers.get("user-id")
    user = Users.objects.filter(id=user_id)
    return user[0]
    # token = request.META.get('HTTP_AUTHORIZATION', False)
    # if token:
    #     token = str(token).split()[1].encode("utf-8")
    #     knoxAuth = TokenAuthentication()
    #     user, auth_token = knoxAuth.authenticate_credentials(token)
    #     request.user = user
    #     return user


def save_file_locally(file):
    """
    Save the uploaded file to a local directory temporarily.
    :param file: The file to be saved locally
    :return: The local file path or None if saving fails
    """
    try:
        # Create a temporary file path
        # local_file_path = os.path.join("/home/ubuntu167/HDD/Projects/Kumar/Project/local-be/cmt-dev-be-python/cmtbackend/tempfiles", file.name)
        directory_path = config('TEMP_FILES_PATH', default='/tmp/ryd_uploads')

        # Check if the directory exists, if not, create it
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)
        local_file_path = os.path.join(directory_path, file.name)
        
        # Open the local file in write-binary mode and save the content
        with open(local_file_path, 'wb') as local_file:
            for chunk in file.chunks():
                local_file.write(chunk)

        return local_file_path
    except Exception as e:
        print(f"Error saving file locally: {str(e)}")
        return None

def remove_local_file(file_path):
    """
    Remove the local file after upload to S3.
    :param file_path: The path of the local file to be deleted
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"Local file {file_path} removed.")
    except Exception as e:
        print(f"Error removing local file: {str(e)}")

# API View for file upload
# class FileUploadAPIView(APIView):
def reusable_file_upload(user, file=None, request_data={}, original_filename=None, is_upload=True):
    """
    Reusable function to handle file upload, metadata management, and database entry.

    :param file: The file to upload
    :param request_data: The request data containing additional metadata (e.g., description, module_name)
    :param is_upload: Flag to determine if file should be uploaded to S3 (default is True)
    :return: Response with file upload status and metadata
    """
    try:
        print("here is reusable code")
        # Ensure file is provided
        if not file:
            print("Inside the if for not file provided")
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        

        print("Got the file")
        logger.info("Processing Job reusable")
        # Default `is_upload` to True if not provided in request
        upload_url = None
        module_name = request_data.get('module_name', None)
        bucket_name = request_data.get("bucket_name", config("AWS_STORAGE_BUCKET_NAME"))
        
        if not original_filename:
            filename = file.name
        else: 
            filename = original_filename.split('/')[-1]
            user_id_str = filename.split("_", 1)[0]
            filename = filename.split("_", 1)[-1]
            try:
                user_id = int(user_id_str)
                user = Users.objects.get(id=user_id)
            except (ValueError, Users.DoesNotExist):
                user = None  
        # Handle file upload to S3 if `is_upload` is True
        if not is_upload:
            # Save file locally first
            local_file_path = save_file_locally(file) if not original_filename else original_filename 
            if not local_file_path:
                return JsonResponse({"error": "Failed to save file locally"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            try:
                print("File : ", file)
                print("File Name : ", filename)
                file_name = f"{module_name}/{filename}"
                
                # Now upload the file to S3
                upload_url = upload_file_to_s3(file, local_file_path, bucket_name, file_name)
                if not original_filename:
                    remove_local_file(local_file_path)
                print("upload url is ------> ", upload_url)
            except Exception as e:
                print(f"Error uploading file to S3: {str(e)}")
                if not original_filename:
                    remove_local_file(local_file_path)
                return JsonResponse({"error": "File upload to S3 failed" + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            upload_url = request_data.get('upload_url', None)

        # Count the records in the file (this can be adjusted based on the actual file structure)
        record_count = count_records_in_file(file)
        print("record count ------> ",record_count)

        # Prepare the file info for database storage
        file_info = {
            "upload_url": file_name,
            "record_count": record_count,
        }
        
        print("file info", file_info)
        try:
            # Store the file record in the database
            file_record = FileManagement.objects.create(
                module_name=module_name,
                file_name=filename,
                # file_size=file.size,
                file_size=file.size if not original_filename else os.path.getsize(original_filename),
                file_path=file_name,
                file_type=filename.split('.')[-1],  # Extract file type based on the extension
                file_info=json.dumps(file_info),  # Store the file info as a JSON string
                created_by=user  # Assuming request_data has the user (adjust if needed)
            )
        except Exception as e:
            print("file management error ------> ",e)
            

        # Return a successful response with the file record details
        return Response({
            "file_record_id": file_record.id,
            "file_name": filename,
            "upload_url": upload_url,
            "record_count": record_count,
            "message": "File uploaded successfully"
        }, status=status.HTTP_201_CREATED)

    except ValidationError as e:
        # Handle validation errors
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        # General exception handling
        return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["PATCH"])
def update_archive(request):
    try:
        file_ids = request.data.get('file_ids', [])
        FileManagement.objects.filter(id__in=file_ids).update(archived=True)
        return Response({"message": "Files archived successfully"})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(["GET"])
def file_download(request):
    """Generate a pre-signed URL for file download and stream the file content."""
    try:
        # Get and validate request parameters
        file_name = request.GET.get("fileName")
        module_name = request.GET.get("moduleName")
        bucket_name = config("AWS_STORAGE_BUCKET_NAME")

        if not file_name:
            return JsonResponse(
                {"error": "fileName query param is required"}, 
                status=400
            )
        
        if not bucket_name:
            return JsonResponse(
                {"error": "bucketName query param is required"}, 
                status=400
            )


        file_name = f"{module_name}/{file_name}"

        # URL encode the filename to handle special characters
        encoded_filename = urllib.parse.quote(file_name)
        
        # Initialize S3 client
        try:
            s3_client = boto3.client(
                "s3",
                aws_access_key_id=config("AWS_ACCESS_KEY"),
                aws_secret_access_key=config("AWS_SECRET_KEY"),
                region_name=config("AWS_S3_REGION_NAME"),
            )
        except Exception as e:
            logger.error(f"Failed to initialize S3 client: {str(e)}")
            return JsonResponse(
                {"error": "Failed to initialize S3 connection"}, 
                status=500
            )

        # Generate presigned URL
        try:
            url = s3_client.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": bucket_name,
                    "Key": file_name,  # Using original file_name as the key
                    "ResponseContentDisposition": f'inline; filename="{encoded_filename}"',
                    "ResponseContentType": "application/octet-stream"
                },
                ExpiresIn=60,  # URL valid for 1 minute
            )
        except Exception as e:
            logger.error(f"Failed to generate presigned URL: {str(e)}")
            return JsonResponse(
                {"error": "Failed to generate download URL"}, 
                status=500
            )
        
        # Make the GET request to the S3 URL
        try:
            s3_response = requests.get(url, stream=True)
            s3_response.raise_for_status()
        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to fetch file from S3: {str(e)}")
            return JsonResponse(
                {"error": "Failed to fetch file from storage"}, 
                status=500
            )

        # Create response using HttpResponse
        response = HttpResponse()
        
        # Copy content type from S3 response
        response['Content-Type'] = s3_response.headers.get('Content-Type', 'application/octet-stream')
        
        # Set content disposition
        response['Content-Disposition'] = f'inline; filename="{encoded_filename}"'
        
        # Copy content length if available
        if 'Content-Length' in s3_response.headers:
            response['Content-Length'] = s3_response.headers['Content-Length']

        # Stream the content
        for chunk in s3_response.iter_content(chunk_size=8192):
            response.write(chunk)

        return response

    except Exception as e:
        logger.error(f"Unexpected error in file_download: {str(e)}")
        return JsonResponse(
            {"error": "An unexpected error occurred"}, 
            status=500
        )
    

def archive_files(filters):
    try:
        if "ids" in filters:
            file_ids = filters['ids']
            FileManagement.objects.filter(id__in=file_ids).update(archived=True)
        elif "file_names" in filters:
            file_names = filters['file_names']
            FileManagement.objects.filter(file_name__in=file_names).update(archived=True)
        return Response({"message": "Files archived successfully"})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class FileDownloadAPIView(APIView):
    """
    API View to download a file.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, file_id, *args, **kwargs):
        try:
            file = FileManagement.objects.get(id=file_id)
        except FileManagement.DoesNotExist:
            return Response({"error": "File not found"}, status=status.HTTP_404_NOT_FOUND)

        file_path = file.file_path

        if file_path.startswith('s3://'):
            # If the file is on S3, fetch it
            s3 = boto3.client('s3')
            bucket_name = settings.AWS_STORAGE_BUCKET_NAME
            file_content = s3.get_object(Bucket=bucket_name, Key=file_path[5:])
            return HttpResponse(file_content['Body'].read(), content_type='application/octet-stream', status=200)
        else:
            # If the file is local
            file_path = os.path.join(settings.MEDIA_ROOT, file_path)
            return HttpResponse(open(file_path, 'rb').read(), content_type='application/octet-stream', status=200)
        
@api_view(["GET"])
def downloadFiles(request):
    if request.method == "GET":
        """
        Download files as xlsx file
        """
        try:
            queryset = FileManagement.objects.all().order_by("-created_at")

            # Convert queryset to list of dictionaries
            file_data = []
            for record in queryset:
                file_data.append(
                    {
                        "id": record.id,
                        "File Name": record.file_name,
                        "File Path": record.file_path,
                        "File Type": record.file_type,
                        "File Size": record.file_size,
                        "File Info": record.file_info,
                        "Module Name": record.module_name,
                        "Archived": record.archived,
                        "Created By": record.created_by,
                        "Created At": record.created_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.created_at else "",
                        "Updated At": record.updated_at.strftime("%d-%m-%Y %H:%M:%S")
                        if record.updated_at else ""
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(file_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Files")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Files"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=file_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


