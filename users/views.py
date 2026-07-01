import json
import logging
import random
import string
from datetime import datetime, timezone, date
from io import BytesIO

import pandas as pd
import pyotp
from decouple import config
from django.conf import settings
from django.contrib.auth import logout
from django.contrib.auth.hashers import check_password, make_password
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django_otp.plugins.otp_totp.models import TOTPDevice
from documents.models import *
from documents.utils.encryption_util import decrypt_text
from helpers.azure_token_helper import (
    decode_azure_token,
    generate_azure_token,
)
from knox.auth import TokenAuthentication
from knox.models import AuthToken
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rest_framework import generics, status, viewsets
from rest_framework.decorators import (
    action,
    api_view,
    permission_classes,
    renderer_classes,
)
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from backend.token_verification_middleware import LoginMethods
from .constants import UserRole

from .models import CoachCoachee, ResetTokens, UserPermissions, Users, UserSessionManagement
from .serializers import (
    CoachCoacheeListSerializer,
    CoachCreateCoacheeSerializer,
    CoachLinkCoacheeSerializer,
    CoachListItemSerializer,
    CoachUpdateCoacheeSerializer,
    LoginSerializer,
    RegisterSerializer,
    UserPermissionsSerializer,
    UserProfileSerializer,
    UserSerializer,
)
from .utils import generate_combination, send_email, send_welcome_password_email
from .auth_helpers import build_user_payload, find_user_by_email, set_auth_cookies
from documents.utils.encryption_util import encrypt_text

logger = logging.getLogger(__name__)

datetime = datetime.datetime

def get_user(request):
    user_id = request.headers.get("user-id")
    user = Users.objects.filter(id=user_id)
    return user[0]

class CountModelMixin(object):
    @action(detail=False)
    def count(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        content = {"count": queryset.count()}
        return Response(content)


# @login_required
class UserViewSet(viewsets.ModelViewSet, CountModelMixin):
    model = Users
    serializer_class = UserSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    # permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def download_users_xlsx(self, request):
        """
        Download user data as xlsx file
        """
        try:
            queryset = self.get_queryset().order_by("-created_at")

            # Convert queryset to list of dictionaries
            user_data = []
            for user in queryset:
                user_data.append(
                    {
                        "id": user.id,
                        "User name": user.user_name,
                        "Email": user.email,
                        "Role": user.role,
                        "Status": user.status,
                        "Department": user.department,
                        "Reporting Manager": user.reporting_manager.user_name,
                        "Division": user.division,
                        "Created at": user.created_at.strftime("%d-%m-%Y %H:%M:%S")
                        if user.created_at
                        else "",
                        "User start date": user.user_start_date.strftime(
                            "%d-%m-%Y %H:%M:%S"
                        )
                        if user.user_start_date
                        else "",
                        "User end date": user.user_end_date.strftime(
                            "%d-%m-%Y %H:%M:%S"
                        )
                        if user.user_end_date
                        else "",
                        "Last login": user.last_login.strftime("%d-%m-%Y %H:%M:%S")
                        if user.last_login
                        else "",
                    }
                )

            # Create DataFrame and Excel file
            df = pd.DataFrame(user_data)
            excel_file = BytesIO()
            df.to_excel(excel_file, index=False, sheet_name="Users")

            # Prepare response
            excel_file.seek(0)
            workbook = load_workbook(excel_file)
            worksheet = workbook["Users"]

            # Define the color for the header row (Color code: #ffc619, RGB: (255, 198, 25))
            header_fill = PatternFill(
                start_color="ffc619", end_color="ffc619", fill_type="solid"
            )

            # Apply the fill to the header row (the first row)
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Save the modified Excel file back to the BytesIO stream
            modified_excel_file = BytesIO()
            workbook.save(modified_excel_file)
            modified_excel_file.seek(0)

            response = HttpResponse(
                modified_excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=users_data.xlsx"
            return response

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get_queryset(self):
        user = Users.objects.all()
        return user

    def list(self, request):
        user = Users.objects.all()
        user_status = request.query_params.get("status", None)
        filter_conditions = Q()
        if user_status:
            filter_conditions &= Q(status=user_status)

        filtered_data = user.filter(filter_conditions)

        serializer = UserSerializer(filtered_data, many=True)
        data = serializer.data
        return Response(data)

    def retrieve(self, request, pk=None):
        if pk:
            user = Users.objects.get(id=pk)
            serializer = UserSerializer(user)
            data = serializer.data
            return Response(data)

    def create(self, request, *args, **kwargs):
        length = 8
        all = string.ascii_letters + string.digits + string.punctuation
        passwords = "".join(random.sample(all, length))
        data = request.data
        user_per = UserPermissions.objects.filter(role=data["role"]).last()
        id = user_per.id
        user_per_id = UserPermissions.objects.get(id=id)
        user = Users.objects.filter(email=data["email"].lower())
        report_m = Users.objects.get(id=data["reporting_manager"])
        division = Entity.objects.get(id=data["division"])
        user_start_date = None
        user_end_date = None
        if data["status"] == "Active":
            user_start_date = datetime.today()
        elif data["status"] == "Inactive":
            user_end_date = datetime.today()
        if user:
            return HttpResponse(
                json.dumps({"message": "User email already exists"}),
                content_type="application/json",
            )
        passwords = generate_combination()
        password = make_password(passwords)
        new_user = Users.objects.create(
            user_name=data["user_name"],
            password=password,
            email=data["email"].lower(),
            role=data["role"],
            status=data["status"],
            reporting_manager=report_m,
            division=division,
            department=data["department"],
            user_permissions=user_per_id,
            user_start_date=user_start_date,
            user_end_date=user_end_date,
        )

        # new_user = Users.objects.create(user_name=data["user_name"], password=password,
        #                                 email=data["email"], role=data["role"],
        #                                 status=data["status"], user_permissions=user_per_id)
        new_user.save()
        if new_user:
            send_welcome_password_email(
                data["email"].lower(),
                passwords,
                subject="Welcome to Mosaic Insurance",
            )

        serializer = UserSerializer(new_user)
        data = serializer.data
        return Response(data)

    def update(self, request, *args, **kwargs):
        user_object = self.get_object()
        data = request.data
        user_object.user_name = data["user_name"]
        user_object.email = data["email"].lower()
        user_object.role = data["role"]
        user_object.status = data["status"]
        user_object.department = data["department"]

        report_m = Users.objects.get(id=data["reporting_manager"]["id"])

        user_object.reporting_manager = report_m

        division = Entity.objects.get(id=data["division"]["id"])

        user_object.division = division

        user_per = UserPermissions.objects.filter(role=data["role"]).last()

        id = user_per.id

        user_per_id = UserPermissions.objects.get(id=id)
        user_object.user_permissions = user_per_id

        if data["status"] == "Active":
            user_object.user_start_date = datetime.today()
        elif data["status"] == "Inactive":
            user_object.user_end_date = datetime.today()

        user_object.save()
        serializer = UserSerializer(user_object)
        data = serializer.data
        return Response(data)

    def partial_update(self, request, *args, **kwargs):
        user_object = self.get_object()
        data = request.data

        try:
            audit_obj = None
            # Save user audit
            for key, value in dict(data).items():
                if key not in ['id']:
                    if key == "division":
                        value = value['entity_divisions']
                    elif key == "reporting_manager":
                        value = value['user_name']
                    elif key == "email":
                        value = encrypt_text(value)

                    audit_obj = CommonAudit.objects.create(
                        table_name="users",
                        record_id=user_object.pk,
                        field_name=key,
                        old_value=str(getattr(user_object, key, "")),
                        new_value=value,
                        previous_time=user_object.updated_at if user_object.updated_at else None,
                        current_time=datetime.now(),
                        changed_by=get_user(request),
                        event_type="edit"
                    )

            try:
                report_m = Users.objects.get(id=data["reporting_manager"]["id"])
                user_object.reporting_manager = report_m
            except:
                pass

            try:
                division = Entity.objects.get(id=data["division"]["id"])
                user_object.division = division
            except:
                pass

            try:
                user_per = UserPermissions.objects.filter(role=data["role"]).last()

                id = user_per.id

                user_per_id = UserPermissions.objects.get(id=id)
                user_object.user_permissions = user_per_id

            except:
                pass

            user_object.user_name = data.get("user_name", user_object.user_name)
            user_object.email = (
                data["email"].lower() if data.get("email") else user_object.email
            )
            user_object.role = data.get("role", user_object.role)
            user_object.status = data.get("status", user_object.status)
            user_object.department = data.get("department", user_object.department)
            if data.get('status') and data['status'] == "Inactive":
                user_object.user_end_date = datetime.now()
            if data.get('status') and data['status'] == "Active":
                user_object.user_end_date = None
            user_object.save()
            serializer = UserSerializer(user_object)
            serializer_data = serializer.data
            return Response(serializer_data)
        except Exception as e:
            if audit_obj:
                audit_obj.delete()
            return Response({'error': f'Unexpected error occurs! {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        user_object = self.get_object()
        user_object.delete()
        return Response({"message": "User deleted successfully"})


class LoginApi(generics.GenericAPIView):
    serializer_class = LoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)

        if not serializer.is_valid():
            non_field_error = serializer.errors.get("non_field_errors")
            field_error = next(iter(serializer.errors.values()), [None])[0]
            message = non_field_error[0] if non_field_error else field_error or "Invalid credentials"
            return Response(
                {"message": message, "success": False, "statusCode": 400},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = serializer.validated_data["user"]
        user.last_login = timezone.now()
        user.save(update_fields=["last_login"])

        response = JsonResponse(
            {
                "message": "Login successful",
                "success": True,
                "user": build_user_payload(user),
            },
            status=200,
        )
        set_auth_cookies(response, user)
        return response


@csrf_exempt
@api_view(["POST"])
def check_login(request):
    """First-time login: email only — generates and emails a password. Returning users get password field."""
    email = (request.data.get("email") or "").lower().strip()
    if not email:
        return Response(
            {"message": "Email is required.", "success": False, "statusCode": 400},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = find_user_by_email(email, require_active=True)
    if not user:
        inactive_user = find_user_by_email(email, require_active=False)
        if inactive_user and inactive_user.status == "Inactive":
            return Response(
                {"message": "Your account is inactive. Please contact support.", "success": False, "statusCode": 400},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(
            {"message": "No account found with this email address.", "success": False, "statusCode": 400},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if user.last_login is None:
        new_password = generate_combination()
        user.password = make_password(new_password)
        user.save(update_fields=["password"])
        send_welcome_password_email(
            decrypt_text(user.email),
            new_password,
            subject="Your login password",
        )
        return JsonResponse(
            {
                "message": "A login password has been sent to your email. Use it to sign in.",
                "success": True,
                "passwordSent": True,
                "showPasswordField": True,
                "statusCode": 200,
            },
            status=status.HTTP_200_OK,
        )

    return JsonResponse(
        {
            "message": "Enter your password to continue.",
            "success": True,
            "passwordSent": False,
            "showPasswordField": True,
            "statusCode": 200,
        },
        status=status.HTTP_200_OK,
    )


class RegisterApi(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)

        if not serializer.is_valid():
            first_error = next(iter(serializer.errors.values()), ["Registration failed"])[0]
            return Response(
                {"message": first_error, "errors": serializer.errors, "success": False},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = serializer.save()
        return JsonResponse(
            {
                "message": "Account created. A login password has been sent to your email.",
                "success": True,
            },
            status=status.HTTP_201_CREATED,
        )


@csrf_exempt
@api_view(["POST"])
def forgotPassword(request):
    model = Users
    serializer_class = UserSerializer
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        email = json_data["email"].lower()
        user = Users.objects.get(email__iexact=email)
        if user:
            newPassword = generate_combination()
            getOtpModel = ResetTokens.objects.filter(user=user)
            if getOtpModel.exists():
                newPassword = getOtpModel[0].resetToken

            else:
                ResetTokens.objects.create(user=user, resetToken=newPassword)
            oldUser = Users.objects.get(id=user.id)
            oldUser.password = make_password(newPassword)
            oldUser.save()
            email_from = settings.EMAIL_HOST_USER
            email_to = email
            recipient_list = [email_to]
            subject = "New Password"
            body = "Your new Password is {} \n\n ENV: {}".format(newPassword, settings.ENVIRONMENT)

            send_email(
                sender_email=email_from,
                recipient_email=recipient_list,
                subject=subject,
                body=body,
            )

            return JsonResponse(
                {
                    "message": "Your password has been reset. New Password has been sent to your email.",
                    "success": True,
                }
            )
        else:
            return JsonResponse({"message": "No such User found", "success": False})


@csrf_exempt
@api_view(["POST"])
def resetPassword(request):
    model = Users
    serializer_class = UserSerializer
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        user_id = json_data["user_id"]
        oldpassword = json_data["oldpassword"]
        password = json_data["password"]
        repassword = json_data["repassword"]
        user = Users.objects.get(id=user_id)
        pas = user.password
        ps = check_password(oldpassword, pas)
        if ps:
            if password == repassword:
                if user:
                    user.password = make_password(password)
                    user.password_updated_at = date.today()
                    user.save()
                    sendResponse = {
                        "message": "Your password has been updated Successfully",
                        "success": True,
                    }
                    return JsonResponse(sendResponse, status=status.HTTP_200_OK)
                else:
                    sendResponse = {
                        "message": "User with this email doesn't exist",
                        "success": False,
                    }
                    return JsonResponse(sendResponse, status=status.HTTP_404_NOT_FOUND)
            else:
                sendResponse = {"message": "Password Do not Match", "success": False}
            return JsonResponse(sendResponse, status=status.HTTP_400_BAD_REQUEST)
        else:
            sendResponse = {"message": "Old Password Do not Match", "success": False}
        return JsonResponse(sendResponse, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(["POST"])
# @login_required
def checkUser(request):
    model = Users
    serializer_class = UserSerializer
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        payload_email = json_data["email"].lower()
        # Use optimized user lookup - first try active users
        user = find_user_by_email(payload_email, require_active=True)

        if not user:
            # Check if user exists but is inactive
            inactive_user = find_user_by_email(payload_email, require_active=False)
            if inactive_user and inactive_user.status == "Inactive":
                sendResponse = {
                    "message": "User is not active!",
                    "success": False,
                    "statusCode": 400,
                }
                return Response(sendResponse, status=400)

            # User doesn't exist at all
            sendResponse = {
                "message": "No account found with the given email address!",
                "success": False,
                "statusCode": 400,
            }
            return Response(sendResponse, status=400)

        if user:
            secret_key = pyotp.random_base32()

            # otp = pyotp.TOTP(secret_key, interval=300)

            otp = pyotp.TOTP(secret_key, interval=1200)

            otpObjects = TOTPDevice.objects.filter(user=user, confirmed=True)
            if otpObjects.exists():
                otpObjects.delete()
            totp_device = TOTPDevice.objects.create(user=user, confirmed=True)
            otp_code = otp.now()
            totp_device.key = secret_key
            totp_device.save()

            email_from = settings.EMAIL_HOST_USER
            email_to = decrypt_text(user.email)
            recipient_list = [email_to]
            subject = "OTP for password reset"
            body = (
                "Your OTP is: "
                + str(otp_code)
                + "\n. Use this OTP to set new password!"
                + "\n\n ENV: " + settings.ENVIRONMENT
            )
            send_email(email_from, recipient_list, subject, body)
            sendResponse = {
                "Message": "OTP is send, please check your mail.",
                "statusCode": status.HTTP_200_OK,
            }
            return JsonResponse(sendResponse, status=status.HTTP_200_OK)
        else:
            sendResponse = {
                "message": "User with this email doesn't exist",
                "success": False,
                "statusCode": status.HTTP_404_NOT_FOUND,
            }

        return JsonResponse(sendResponse, status=status.HTTP_404_NOT_FOUND)


@csrf_exempt
@api_view(["POST"])
# @login_required
def verifyOTP(request):
    model = Users
    serializer_class = UserSerializer
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        payload_email = json_data["email"].lower()
        entered_otp = json_data["otp"]
        # Use optimized user lookup - first try active users
        user = find_user_by_email(payload_email, require_active=True)

        if not user:
            # Check if user exists but is inactive
            inactive_user = find_user_by_email(payload_email, require_active=False)
            if inactive_user and inactive_user.status == "Inactive":
                sendResponse = {
                    "message": "User is not active!",
                    "success": False,
                    "statusCode": 400,
                }
                return Response(sendResponse, status=400)

            # User doesn't exist at all
            sendResponse = {
                "message": "No account found with the given email address!",
                "success": False,
                "statusCode": 400,
            }
            return Response(sendResponse, status=400)

        if user:
            try:
                totp_device = TOTPDevice.objects.filter(user=user).last()
                otp = pyotp.TOTP(totp_device.key, interval=1200)

                if otp.verify(entered_otp):
                    TOTPDevice.objects.filter(user=user).delete()
                    sendResponse = {
                        "message": "OTP has been successfully verified!",
                        "success": True,
                        "statusCode": 200,
                    }
                    return JsonResponse(sendResponse, status=status.HTTP_200_OK)
                else:
                    sendResponse = {
                        "message": "Invalid or expired OTP. Please try again.",
                        "success": False,
                        "statusCode": 400,
                    }
                    return JsonResponse(sendResponse, status=status.HTTP_200_OK)
            except:
                sendResponse = {
                    "message": "Please try again with OTP generation & verification!",
                    "success": False,
                    "statusCode": 400,
                }
                return JsonResponse(sendResponse, status=status.HTTP_200_OK)
        else:
            sendResponse = {
                "message": "User with this email doesn't exist",
                "success": False,
                "statusCode": 400,
            }
        return JsonResponse(sendResponse, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["POST"])
# @login_required
def createPassword(request):
    model = Users
    serializer_class = UserSerializer
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        payload_email = json_data["email"].lower()
        password = json_data["password"]
        repassword = json_data["repassword"]
        # Use optimized user lookup - first try active users
        user = find_user_by_email(payload_email, require_active=True)

        if not user:
            # Check if user exists but is inactive
            inactive_user = find_user_by_email(payload_email, require_active=False)
            if inactive_user and inactive_user.status == "Inactive":
                sendResponse = {
                    "message": "User is not active!",
                    "success": False,
                    "statusCode": 400,
                }
                return Response(sendResponse, status=400)

            # User doesn't exist at all
            sendResponse = {
                "message": "No account found with the given email address!",
                "success": False,
                "statusCode": 400,
            }
            return Response(sendResponse, status=400)

        if password == repassword:
            if user:
                user.password = make_password(password)
                user.password_updated_at = date.today()
                user.save()
                sendResponse = {
                    "message": "Your password has been updated Successfully!",
                    "success": True,
                    "statusCode": 200,
                }
                return JsonResponse(sendResponse, status=status.HTTP_200_OK)
            else:
                sendResponse = {
                    "message": "User with this email doesn't exist",
                    "success": False,
                    "statusCode": 400,
                }
                return JsonResponse(sendResponse, status=status.HTTP_404_NOT_FOUND)
        else:
            sendResponse = {
                "message": "Password do not Match",
                "success": False,
                "statusCode": 400,
            }
        return JsonResponse(sendResponse, status=status.HTTP_400_BAD_REQUEST)


class UserPermissionViewSet(viewsets.ModelViewSet):
    queryset = UserPermissions.objects.all().order_by("-id")
    serializer_class = UserPermissionsSerializer
    
    def create(self, request, *args, **kwargs):
        role_options = [
            'Add',
            'Edit',
            'View',
            'Delete'
        ]
        data=request.data
        adding_module = data['addModule']
        if adding_module:
            # add module process
            try:
                add_module_data = {data['module']: {}}
                obj = self.queryset.get(role=data['role'])
                if data['module'] in list(obj.permissions_list.keys()):
                    return Response({"error": "Module already exists with same role!"}, status=status.HTTP_400_BAD_REQUEST)
                for i in data['permissions']:
                    add_module_data[data['module']].update({i: 'Y'})
                for j in role_options:
                    if j not in data['permissions']:
                        add_module_data[data['module']].update({j: 'N'})
                old_module_data = obj.permissions_list
                old_module_data.update(add_module_data)
                obj.permissions_list = old_module_data
                obj.save()
                return Response({"success": "Module added successfully."}, status=200)
            except Exception as e:
                return Response({"error": "Module addtion falied!"}, status=status.HTTP_400_BAD_REQUEST)
        elif self.queryset.filter(role=data['role']).exists():
            return Response({"error": "Role already exists!"}, status=status.HTTP_400_BAD_REQUEST)

        # add role process
        input_data = {}
        input_data['role'] = data['role']
        input_data['permissions_list'] = {data['module']: {}}
        for i in data['permissions']:
            input_data['permissions_list'][data['module']].update({i: 'Y'})
        for j in role_options:
            if j not in data['permissions']:
                input_data['permissions_list'][data['module']].update({j: 'N'})
        serializer = self.serializer_class(data=input_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=201)

    def list(self, request):
        user_status = request.query_params.get("status", None)
        filter_conditions = Q()
        if user_status:
            filter_conditions &= Q(status=user_status)

        filtered_data = self.queryset.filter(filter_conditions)

        serializer = self.serializer_class(filtered_data, many=True)
        data = serializer.data
        return Response(data)

    def partial_update(self, request, *args, **kwargs):
        data = request.data
        permission_object = self.get_object()

        try:
            module = data.get("selectedModule")
            permission_key = data.get("permission_key")
            value = data.get("value")
            print('...?',module, permission_key, value)
            permissions = permission_object.permissions_list

            if module in permissions:
                if permission_key in permissions[module]:
                    permissions[module][permission_key] = value
                    # Save object
                    permission_object.permissions_list = permissions
                    permission_object.save()

                else:
                    return Response({"error": "Permission key not found"}, status=400)
            else:
                return Response({"error": "Module not found"}, status=400)

            serializer = self.serializer_class(permission_object)
            return Response(serializer.data)

        except Exception as e:
            return Response(
                {"error": f"Unexpected error occurs! {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # def destroy(self, request, *args, **kwargs):
    #     user_object = self.get_object()
    #     user_object.delete()
    #     return Response({"message": "User deleted successfully"})


@api_view(["POST"])
def user_logout(request):
    try:
        user_id = request.data.get("user_id")

        if not user_id:
            return JsonResponse(
                {"message": "User ID is required.", "success": False},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = Users.objects.get(id=user_id)
        except Users.DoesNotExist:
            return JsonResponse(
                {"message": "User not found.", "success": False},
                status=status.HTTP_404_NOT_FOUND
            )

        # Clear all cookies
        response = JsonResponse(
            {"message": "User successfully logged out.", "success": True},
            status=status.HTTP_200_OK
        )

        for cookie in request.COOKIES:
            response.delete_cookie(cookie)

        # Update UserSessionManagement
        UserSessionManagement.objects.update_or_create(
            user=user,
            defaults={
                "logged_in_time": None,
                "isLoggedIn": False
            },
        )
        logger.info("UserSessionManagement updated, User logged out: %s", user.email)

        return response

    except Exception as e:
        logger.exception("Logout failed.")
        return JsonResponse(
            {"message": "Logout failed.", "error": str(e), "success": False},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["POST"])
def deactivateUser(request):
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        email = json_data["email"].lower()

        user = Users.objects.filter(email__iexact=email).last()
        if user:
            user.status = "Inactive"
            user.user_end_date = datetime.today()
            user.save()
            sendResponse = {
                "message": "User has been deactivated Successfully!",
                "success": True,
                "statusCode": 201,
            }
            return JsonResponse(sendResponse, status=status.HTTP_200_OK)
        else:
            sendResponse = {
                "message": "User with this email doesn't exist",
                "success": False,
                "statusCode": 400,
            }
        return JsonResponse(sendResponse, status=status.HTTP_404_NOT_FOUND)


@api_view(["POST"])
def activateUser(request):
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        email = json_data["email"].lower()

        user = Users.objects.filter(email__iexact=email).last()
        if user:
            passwords = generate_combination()
            user.password = make_password(passwords)
            user.status = "Active"
            user.user_start_date = datetime.today()
            user.save()

            send_welcome_password_email(email, passwords, subject="Welcome to Mosaic Insurance")
            sendResponse = {
                "Message": "Email is sent successfully!",
                "statusCode": status.HTTP_200_OK,
            }
            return JsonResponse(sendResponse, status=status.HTTP_200_OK)
        else:
            sendResponse = {
                "message": "User with this email doesn't exist",
                "success": False,
                "statusCode": 400,
            }
        return JsonResponse(sendResponse, status=status.HTTP_404_NOT_FOUND)


@api_view(["POST"])
def checkOldPassword(request):
    if request.method == "POST":
        json_data = json.loads(request.body.decode("utf-8"))
        email = json_data["email"].lower()
        pwd = json_data["password"]

        user = Users.objects.filter(email__iexact=email).last()
        isChecked = check_password(pwd, user.password)
        if isChecked:
            sendResponse = {
                "message": "Password matched!",
                "success": True,
                "statusCode": 200,
            }
            return JsonResponse(sendResponse, status=status.HTTP_200_OK)
        else:
            sendResponse = {
                "message": "Password not matched!",
                "success": False,
                "statusCode": 400,
            }
        return JsonResponse(sendResponse, status=status.HTTP_400_BAD_REQUEST)


class GenerateNewAccessToken(generics.GenericAPIView):
    """
    This class for generate new access using refresh token
    """

    def post(self, request):
        """
        @param request: it takes request as argument with refresh token in header
        @return: it return access token after verify refresh token
        """

        token = request.data.get("refresh", None)
        try:
            if token:
                token_authentication = TokenAuthentication()
                user, token_model = token_authentication.authenticate_credentials(
                    token.encode()
                )
                user_operations_obj = Users.objects.get(email=user.email)
                return Response(
                    {
                        "access": AuthToken.objects.create(user_operations_obj)[1],
                        "refresh": AuthToken.objects.create(user_operations_obj)[1],
                    }
                )
        except:
            response = {
                "message": "Please provide the refresh token",
                "success": False,
                "statusCode": 400,
            }
            return Response(response, status=400)


class DecryptUserDataAPIView(APIView):
    def post(self, request):
        data = request.data
        user_id = data.get("id")
        type = data.get("type")
        record_id = data.get("record_id")

        try:
            user_info = Users.objects.get(id=user_id)
        except Users.DoesNotExist:
            return Response(
                {"error": "User with this ID does not exist"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if type == "decrypted_email":
            decrypted_value = user_info.get_decrypted_email()
        elif type == "old_email_value":
            audit_obj = CommonAudit.objects.get(id=record_id)
            decrypted_value = user_info.get_decrypted_value(audit_obj.old_value)
        elif type == "new_email_value":
            audit_obj = CommonAudit.objects.get(id=record_id)
            decrypted_value = user_info.get_decrypted_value(audit_obj.new_value)
        else:
            return Response(
                {"error": "Invalid text type provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        response_data = {type: decrypted_value}
        return Response(response_data, status=status.HTTP_200_OK)

@csrf_exempt
@api_view(["POST"])
def oauth_login(request):
    """Handles Azure AD OAuth login and sets cookies."""

    def set_cookie(response, key, value, max_age):
        response.set_cookie(
            key=key,
            value=value,
            httponly=True,
            secure=True,
            max_age=max_age,
        )

    try:
        data = json.loads(request.body.decode("utf-8"))
        authorization_code = data.get("authorization_code")
        if not authorization_code:
            return JsonResponse(
                {"message": "Authorization code is required", "success": False},
                status=status.HTTP_400_BAD_REQUEST,
            )

        logger.info("Generating Azure AD token")
        token_response = generate_azure_token(authorization_code)
        access_token = token_response["access_token"]
        refresh_token = token_response["refresh_token"]
        logger.info("Azure token generated successfully")

        decoded_token = decode_azure_token(access_token)
        email = decoded_token.get("upn", "").lower()
        logger.info("Decoded user email from token: %s", email)

        user_info = find_user_by_email(email, require_active=True)

        if not user_info:
            inactive_user = find_user_by_email(email, require_active=False)
            if inactive_user and inactive_user.status == "Inactive":
                logger.info("Inactive user tried to login: %s", email)
                return JsonResponse(
                    {
                        "message": "User is not active!",
                        "success": False,
                        "statusCode": 400,
                    },
                    status=400,
                )

            logger.info("User not found: %s", email)
            return JsonResponse(
                {
                    "message": "No account found with the given email address!",
                    "success": False,
                    "statusCode": 400,
                },
                status=400,
            )

        logger.info("User found: %s", user_info.email)

        response_data = {
            "access_token": access_token,
            "username": user_info.user_name,
            "email": user_info.email,
            "role": user_info.role,
            "status": user_info.status,
            "user_id": user_info.id,
            "user": user_info.id,
        }

        response = JsonResponse(response_data, status=200)

        # Set cookies
        cookie_map = {
            "accessToken": access_token,
            "refreshToken": refresh_token,
            "userId": user_info.id,
            "loginMethod": LoginMethods.OAUTH.value,
        }

        for key, val in cookie_map.items():
            set_cookie(response, key, val, 24 * 60 * 60)

        UserSessionManagement.objects.update_or_create(
            user=user_info,
            defaults={
                "logged_in_time": datetime.now(timezone.utc),
                "isLoggedIn": True
            },
        )
        logger.info("UserSessionManagement updated, User logged in: %s", user_info.email)

        # last login update
        user_info.last_login = datetime.now()
        user_info.save()

        return response

    except Exception as e:
        logger.exception("OAuth login failed")
        return JsonResponse(
            {"message": "Server Error", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@api_view(['POST'])
def check_user_exists_or_not(request):
    email = request.data.get('currentEmail')
    if not email:
        return JsonResponse({"message": "Email is required"}, status=status.HTTP_400_BAD_REQUEST)

    # Use values_list to fetch only email field to reduce DB load
    emails = Users.objects.values_list('email', flat=True)

    for encrypted_email in emails:
        try:
            decrypted = decrypt_text(encrypted_email)
            if decrypted.lower() == email.lower():
                return JsonResponse({"message": "Yes"}, status=status.HTTP_200_OK)
        except Exception:
            continue

    return JsonResponse({"message": "No"}, status=status.HTTP_200_OK)


@api_view(['GET'])
def get_users_list(request):
    user = Users.objects.all()
    user_name = request.query_params.get("userName", None)
    user_role = request.query_params.get("userRole", None)
    user_status = request.query_params.get("userStatus", None)

    page_number = int(request.GET.get("skip", 0))
    rows_per_page = int(request.GET.get("pageSize", 20))
    skip = page_number * rows_per_page
    
    filter_conditions = Q()
    if user_name:
        filter_conditions &= Q(user_name__icontains=user_name)
    if user_role:
        filter_conditions &= Q(role__icontains=user_role)
    if user_status:
        filter_conditions &= Q(status=user_status)

    filtered_data = user.filter(filter_conditions).order_by('-id')
    count = len(filtered_data)

    serializer = UserSerializer(filtered_data[skip: skip + rows_per_page], many=True)
    data = serializer.data
    return Response({ "data": data, "count": count }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
def check_auth(request):
    user_id = request.COOKIES.get("userId")
    if not user_id:
        return JsonResponse({"authenticated": False}, status=status.HTTP_200_OK)

    try:
        user = Users.objects.get(id=user_id)
    except Users.DoesNotExist:
        return JsonResponse({"authenticated": False}, status=status.HTTP_200_OK)

    return JsonResponse(
        {
            "authenticated": True,
            "user": build_user_payload(user),
        },
        status=status.HTTP_200_OK,
    )


def _get_coach_from_request(request):
    coach = getattr(request, "user", None)
    if coach and isinstance(coach, Users):
        return coach

    user_id = request.COOKIES.get("userId")
    if not user_id:
        return None

    try:
        return Users.objects.get(id=user_id)
    except Users.DoesNotExist:
        return None


@api_view(["GET", "POST"])
def list_my_coachees(request):
    coach = _get_coach_from_request(request)
    if not coach:
        return Response(
            {"message": "Authentication required.", "success": False},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if coach.role != UserRole.COACH:
        return Response(
            {"message": "Only coaches can manage coachees.", "success": False},
            status=status.HTTP_403_FORBIDDEN,
        )

    if request.method == "POST":
        serializer = CoachCreateCoacheeSerializer(
            data=request.data,
            context={"coach": coach},
        )
        if not serializer.is_valid():
            first_error = next(iter(serializer.errors.values()), ["Invalid data"])[0]
            return Response(
                {
                    "message": first_error,
                    "errors": serializer.errors,
                    "success": False,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        link = serializer.save()
        return Response(
            {
                "message": "Coachee created. A login password has been sent to their email.",
                "success": True,
                "coachee": CoachCoacheeListSerializer(link).data,
            },
            status=status.HTTP_201_CREATED,
        )

    links = (
        CoachCoachee.objects.filter(coach=coach)
        .select_related("coachee")
        .order_by("-created_at")
    )
    serializer = CoachCoacheeListSerializer(links, many=True)
    return Response({"coachees": serializer.data, "count": links.count()})


@api_view(["POST"])
def link_coachee(request):
    coach = _get_coach_from_request(request)
    if not coach:
        return Response(
            {"message": "Authentication required.", "success": False},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if coach.role != UserRole.COACH:
        return Response(
            {"message": "Only coaches can add coachees.", "success": False},
            status=status.HTTP_403_FORBIDDEN,
        )

    serializer = CoachLinkCoacheeSerializer(
        data=request.data,
        context={"coach": coach},
    )
    if not serializer.is_valid():
        first_error = next(iter(serializer.errors.values()), ["Invalid data"])[0]
        return Response(
            {
                "message": first_error,
                "errors": serializer.errors,
                "success": False,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    link = serializer.save()
    return Response(
        {
            "message": "Coachee added to your account successfully.",
            "success": True,
            "coachee": CoachCoacheeListSerializer(link).data,
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(["PATCH", "PUT", "POST"])
def update_my_coachee(request):
    coach = _get_coach_from_request(request)
    if not coach:
        return Response(
            {"message": "Authentication required.", "success": False},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if coach.role != UserRole.COACH:
        return Response(
            {"message": "Only coaches can update coachees.", "success": False},
            status=status.HTTP_403_FORBIDDEN,
        )

    coachee_id = request.data.get("coachee_id")
    if not coachee_id:
        return Response(
            {"message": "coachee_id is required.", "success": False},
            status=status.HTTP_400_BAD_REQUEST,
        )

    link = (
        CoachCoachee.objects.filter(coach=coach, coachee_id=coachee_id)
        .select_related("coachee")
        .first()
    )
    if not link:
        return Response(
            {"message": "Coachee not found on your account.", "success": False},
            status=status.HTTP_404_NOT_FOUND,
        )

    serializer = CoachUpdateCoacheeSerializer(data=request.data, partial=True)
    if not serializer.is_valid():
        first_error = next(iter(serializer.errors.values()), ["Invalid data"])[0]
        return Response(
            {
                "message": first_error,
                "errors": serializer.errors,
                "success": False,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    link.coachee.user_name = serializer.validated_data["user_name"]
    link.coachee.save(update_fields=["user_name"])

    return Response(
        {
            "message": "Coachee name updated successfully.",
            "success": True,
            "coachee": CoachCoacheeListSerializer(link).data,
        },
        status=status.HTTP_200_OK,
    )


def _get_user_from_request(request):
    return _get_coach_from_request(request)


@api_view(["GET"])
def list_coaches(request):
    user = _get_user_from_request(request)
    if not user:
        return Response(
            {"message": "Authentication required.", "success": False},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    coaches = Users.objects.filter(role=UserRole.COACH, status="Active").order_by("user_name")
    serializer = CoachListItemSerializer(coaches, many=True)
    return Response({"coaches": serializer.data, "count": coaches.count()})


@api_view(["GET", "PUT"])
def my_coach(request):
    user = _get_user_from_request(request)
    if not user:
        return Response(
            {"message": "Authentication required.", "success": False},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if user.role != UserRole.COACHEE:
        return Response(
            {"message": "Only coachees can manage their coach.", "success": False},
            status=status.HTTP_403_FORBIDDEN,
        )

    if request.method == "GET":
        coach = user.reporting_manager
        return Response(
            {
                "coach_id": coach.id if coach else None,
                "coach_name": coach.user_name if coach else None,
            }
        )

    coach_id = request.data.get("coach_id")
    if not coach_id:
        return Response(
            {"message": "coach_id is required.", "success": False},
            status=status.HTTP_400_BAD_REQUEST,
        )

    coach = get_object_or_404(Users, id=coach_id, role=UserRole.COACH, status="Active")
    user.reporting_manager = coach
    user.save(update_fields=["reporting_manager"])
    CoachCoachee.objects.filter(coachee=user).exclude(coach=coach).delete()
    CoachCoachee.objects.get_or_create(coach=coach, coachee=user)

    return Response(
        {
            "message": "Coach updated successfully.",
            "success": True,
            "coach_id": coach.id,
            "coach_name": coach.user_name,
            "user": UserProfileSerializer(user).data,
        }
    )
